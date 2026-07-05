# ============================================================================
# Project X
# Logbook XLSX Generator (Duna Monitor compatible)
# ============================================================================

from __future__ import annotations

import csv
from pathlib import Path

from logbook.paths import CSV_FILENAME, XLSX_FILENAME


def regenerate_xlsx(ship_dir: Path) -> Path | None:

    csv_file = ship_dir / CSV_FILENAME
    xlsx_file = ship_dir / XLSX_FILENAME

    if not csv_file.exists():
        return None

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as error:
        raise RuntimeError(
            "openpyxl is required for logbook XLSX generation"
        ) from error

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Duna Monitor"

    with csv_file.open("r", encoding="utf-8") as handle:
        reader = csv.reader(handle, delimiter=";")

        for row in reader:
            worksheet.append(row)

    navy = PatternFill("solid", fgColor="002060")
    dark_navy = PatternFill("solid", fgColor="1F4E78")
    green = PatternFill("solid", fgColor="00B050")
    yellow = PatternFill("solid", fgColor="FFD966")
    orange = PatternFill("solid", fgColor="ED7D31")
    gray = PatternFill("solid", fgColor="7F7F7F")

    white_font = Font(color="FFFFFF", bold=True, size=13)
    data_font = Font(color="FFFFFF", size=11)
    thin = Side(style="thin", color="D9D9D9")

    for cell in worksheet[1]:
        cell.fill = navy
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.row_dimensions[1].height = 34

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.fill = dark_navy
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_index in range(2, worksheet.max_row + 1):
        text = str(worksheet.cell(row_index, 3).value)

        if "Látható" in text and "Talán" not in text:
            worksheet.cell(row_index, 3).fill = green
        elif "Talán" in text:
            worksheet.cell(row_index, 3).fill = yellow
        elif "Nem látható" in text:
            worksheet.cell(row_index, 3).fill = gray

    for row_index in range(2, worksheet.max_row + 1):
        speed_text = str(worksheet.cell(row_index, 5).value)

        try:
            speed = float(speed_text.split()[0])

            if speed <= 0.5:
                worksheet.cell(row_index, 5).fill = green
            elif speed <= 5:
                worksheet.cell(row_index, 5).fill = yellow
            else:
                worksheet.cell(row_index, 5).fill = orange
        except (AttributeError, IndexError, TypeError, ValueError):
            pass

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = str(cell.value or "")

            if len(value) > max_length:
                max_length = len(value)

        worksheet.column_dimensions[column_letter].width = max_length + 4

    worksheet.freeze_panes = "A2"
    workbook.save(xlsx_file)
    return xlsx_file
